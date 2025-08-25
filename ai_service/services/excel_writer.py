import io
import logging
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class ExcelWriter:
    """Service for writing data to Excel files"""
    
    def __init__(self):
        pass
    
    def create_sor_excel(self, items: List[Dict], filename: str = "sor_output.xlsx") -> bytes:
        """
        Create an Excel file with SOR/BOQ data
        
        Args:
            items: List of SOR/BOQ items
            filename: Output filename
            
        Returns:
            Excel file bytes
        """
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "SOR/BOQ"
            
            # Add headers
            headers = ["Item No.", "Description", "Unit", "Quantity", "Rate", "Amount", "Category", "Notes"]
            ws.append(headers)
            
            # Style headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for i, item in enumerate(items, start=1):
                row = [
                    i,  # Item No.
                    item.get("description", ""),
                    item.get("unit", ""),
                    item.get("quantity", ""),
                    item.get("suggested_rate", "") or item.get("rate", ""),
                    item.get("amount", ""),
                    item.get("suggested_category", "") or item.get("category", ""),
                    item.get("notes", "")
                ]
                ws.append(row)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"Created SOR Excel file with {len(items)} items")
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating SOR Excel file: {str(e)}")
            raise
    
    def create_invoice_excel(self, invoice_data: Dict[str, Any], filename: str = "invoice_output.xlsx") -> bytes:
        """
        Create an Excel file with invoice data
        
        Args:
            invoice_data: Invoice data dictionary
            filename: Output filename
            
        Returns:
            Excel file bytes
        """
        try:
            # Create a new workbook
            wb = Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Invoice Summary"
            
            # Add summary data
            ws_summary.append(["Invoice Summary"])
            ws_summary.append([""])
            ws_summary.append(["Invoice Number", invoice_data.get("invoice_number", "")])
            ws_summary.append(["Vendor Name", invoice_data.get("vendor_name", "")])
            ws_summary.append(["Issue Date", invoice_data.get("issue_date", "")])
            ws_summary.append(["Due Date", invoice_data.get("due_date", "")])
            ws_summary.append(["Total Amount", invoice_data.get("total_amount", "")])
            ws_summary.append(["GST Amount", invoice_data.get("gst_amount", "")])
            ws_summary.append(["Subtotal", invoice_data.get("subtotal", "")])
            
            # Style summary title
            ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
            
            # Items sheet
            ws_items = wb.create_sheet("Line Items")
            
            # Add items headers
            item_headers = ["Description", "Quantity", "Unit Price", "Total"]
            ws_items.append(item_headers)
            
            # Style item headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for col in range(1, len(item_headers) + 1):
                cell = ws_items.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add item data
            line_items = invoice_data.get("line_items", [])
            for item in line_items:
                row = [
                    item.get("description", ""),
                    item.get("quantity", ""),
                    item.get("unit_price", ""),
                    item.get("total", "")
                ]
                ws_items.append(row)
            
            # Auto-adjust column widths
            for ws in [ws_summary, ws_items]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info("Created invoice Excel file")
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating invoice Excel file: {str(e)}")
            raise